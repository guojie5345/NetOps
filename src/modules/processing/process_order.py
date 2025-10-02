#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""订单处理模块，负责解析订单文件并生成配置脚本"""

import json
from math import log
import os
import re
import sys
import ipaddress
import jinja2
import requests
import jsonpath
import warnings
from datetime import datetime
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

warnings.filterwarnings('ignore', message="Data Validation extension is not supported and will be removed")

# 动态添加项目根目录到sys.path
current_file = os.path.abspath(__file__)
project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(current_file))))

# 从配置文件中读取项目根目录，如果配置文件中未设置，则使用自动计算的路径
try:
    config_file_path = os.path.join(project_root, 'config', 'config.json')
    if os.path.exists(config_file_path):
        with open(config_file_path, 'r', encoding='utf-8') as f:
            config_data = json.load(f)
            if config_data.get('project', {}).get('root_path'):
                project_root = config_data['project']['root_path']
except Exception as e:
    pass  # 如果读取配置文件出错，继续使用自动计算的路径

if project_root not in sys.path:
    sys.path.insert(0, project_root)

from src.core.config_manager import ConfigManager
from src.utils.logger import setup_logger, get_module_logger

# 定义项目根目录和相关路径
config_path = os.path.join(project_root, 'config', 'config.json')
template_dir = os.path.join(project_root, 'templates')
jsons_dir = os.path.join(project_root, 'data', 'output', 'json_files')

logger = get_module_logger(__name__)


def save_response_to_json(response: object, filename: object, logger=None) -> None:
    """
    将API响应结果保存为 JSON 文件到 json_files 文件夹下。

    :param response: API响应结果（字节流）
    :param filename: API操作名称
    :param logger: 日志记录器实例
    """
    if response is None:
        if logger:
            logger.error("响应数据为空")
        return

    # 确保 json_files 文件夹存在
    if not os.path.exists(jsons_dir):
        os.makedirs(jsons_dir)
    file_path = os.path.join(jsons_dir, f'{filename}')

    try:
        # 类型一：将字节流转换为字符串并解析为JSON
        if isinstance(response, bytes):
            response_str = response.decode('utf-8')
            response_json = json.loads(response_str)
        # 类型二：直接使用 response.json() 方法解析 JSON
        elif isinstance(response, requests.models.Response):
            response_json = response.json()
        # 类型三：直接使用列表
        elif isinstance(response, list):
            response_json = response
        else:
            if logger:
                logger.error(f"响应数据类型不符合保存类型, 当前类型为：{type(response)}")
            return
        # 将JSON数据写入文件
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            json.dump(response_json, f, ensure_ascii=False, indent=4)  # type: ignore
        if logger:
            logger.info(f"响应数据已保存到 {file_path}")
    except Exception as e:
        if logger:
            logger.error(f"保存响应数据到文件时出错: {e}")


def extract_jsonpath(data, path, default=None):
    """
    使用 jsonpath 从数据中提取值，如果未匹配到则返回默认值。

    Args:
        data (dict): 要提取数据的 JSON 对象。
        path (str): jsonpath 表达式。
        default: 未匹配到值时返回的默认值，默认为 None。

    Returns:
        匹配到的值或默认值。
    """
    result = jsonpath.jsonpath(data, path)
    return result[0] if result else default


def _load_json_file(file_path):
    """加载JSON文件并返回数据，包含错误处理"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        logger.error(f"JSON文件未找到: {file_path}")
        raise
    except json.JSONDecodeError:
        logger.error(f"JSON文件解析错误: {file_path}")
        raise


def get_combine_net_infos(ips, vpcs_data):
    """查找IP所在的VPC信息，区分源IP和目的IP"""
    combine_net_infos = {}

    for ip in ips:
        # # 确定IP方向（源或目的）
        # direction = 'source' if ip in s_ips else 'destination'

        # 查找IP所在的阿里云VPC
        for vpc in vpcs_data.get('Vpcs', {}).get('Vpc', []):
            cidr = vpc.get('CidrBlock')
            if not cidr:
                continue
            try:
                ip_obj = ipaddress.ip_interface(ip)
                cidr_obj = ipaddress.ip_interface(cidr)
                if ip_obj.network.subnet_of(cidr_obj.network):
                    logger.info(f"变更中地址： {ip} 在VPC网段 {cidr} 中")
                    combine_net_infos[ip] = {'cidr': cidr_obj.exploded}
                    break  # 找到后跳出循环
            except ValueError as e:
                logger.debug(f"处理ip地址或网段：{ip}不在{cidr}网段中")
                continue
        if not combine_net_infos:
            logger.info(f"IP {ip} 未找到所属VPC")

    return combine_net_infos


class OrderProcessor:
    """订单处理器类，用于处理Excel订单文件并生成网络配置方案
    
    该类负责解析订单文件、提取客户需求、生成配置方案并保存结果。
    支持Excel格式的订单文件，能够处理复杂的网络配置需求。
    """

    def __init__(self, order_path, config, template_dir):
        """初始化订单处理器
        
        :param order_path: str - 订单文件路径
        :param config: dict - 配置信息字典
        :param template_dir: str - 模板目录路径
        """
        self.order_path = order_path
        self.config = config
        self.template_dir = template_dir
        self.logger = get_module_logger(__name__)
        self._process_order()

    def _process_order(self):
        """处理订单文件的主要流程
        
        该方法是订单处理的核心流程，负责协调整个处理过程：
        1. 解析订单文件，提取客户信息和需求详情
        2. 为每个需求生成变更标题、原因和配置方案
        3. 将所有结果保存到JSON文件中
        
        :raises Exception: 当处理过程中发生错误时抛出异常
        """
        results = []
        try:
            self.logger.info(f"开始处理订单文件: {self.order_path}")

            # 解析订单文件
            customer_info, requirement_infos = self.parse_order_file()

            for requirement_info in requirement_infos:
                # 生成变更标题
                title = self.generate_title(customer_info, requirement_info)

                # 生成变更原因
                reason = self.generate_reason(customer_info, requirement_info)

                # 生成变更方案
                scheme = self.generate_scheme(customer_info, requirement_info)

                # 组合返回结果
                result = {
                    "title": title,
                    "reason": reason,
                    "scheme": scheme
                }
                results.append(result)

            # 保存结果到JSON文件
            self._save_results_to_json(results, customer_info)
            
            self.logger.info("订单处理完成")

            return json.dumps(results, ensure_ascii=False)

        except Exception as e:
            self.logger.error(f"订单处理失败: {str(e)}", exc_info=True)
            raise

    def parse_order_file(self):
        """解析订单文件，支持Excel格式
        
        该方法负责解析Excel格式的订单文件，提取客户信息和网络需求详情。
        处理合并单元格的情况，并对IP地址和端口进行标准化处理。
        
        :return: tuple - 客户信息字典和需求信息列表
        """
        # 加载工作簿，设置为可写模式以处理合并单元格
        wb = load_workbook(self.order_path, read_only=False, data_only=True)
        ws = wb.active

        # 用于存储所有需求的列表
        requirements_list = []

        # 从工作表的 F4 单元格获取组织名称
        organization = ws['F4']
        # 从工作表的 F5 单元格获取系统名称
        system_name = ws['F5']
        # 构建包含客户信息的字典
        customer = {"organization": organization.value, "system_name": system_name.value}

        # 获取工作表第 9 行 C 列到 P 列的单元格范围（标题行）
        tts = ws['C9:P9']
        # 使用列表推导式将单元格的值展平为一个列表，并移除值中的星号
        tt_list = [cell.value.replace('*', '') for row in tts for cell in row if cell.value is not None]

        # 从第 10 行开始遍历到工作表的最后一行（数据行）
        for row_num in range(10, ws.max_row + 1):
            # 复制合并单元格范围列表，避免在遍历过程中修改集合导致错误
            merged_ranges = list(ws.merged_cells.ranges)
            # 遍历所有合并单元格范围
            for merged_range in merged_ranges:
                # 检查当前行是否在合并单元格范围内
                if merged_range.min_row <= row_num <= merged_range.max_row:
                    # 获取合并单元格左上角的单元格
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    # 获取合并单元格的值
                    value = top_left_cell.value
                    # 取消当前的合并单元格
                    ws.unmerge_cells(range_string=str(merged_range))
                    # 将合并单元格的值赋给分开后的每个单元格
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            ws.cell(row=row, column=col, value=value)

            # 获取当前行 C 列到 R 列的单元格范围
            values = ws[f'C{row_num}:R{row_num}']
            # 使用列表推导式将单元格的值展平为一个列表，保留空值为 None
            values_list = [cell.value for row in values for cell in row]
            if values_list[0] is None:
                continue
            # 将标题列表和当前行的值列表合并为字典
            d = dict(zip(tt_list, values_list))
            # 对IP地址和端口字段进行标准化处理
            d['源IP'] = self.extra_ip(d['源IP'])
            d['目的IP'] = self.extra_ip(d['目的IP'])
            d['公网IP'] = self.extra_ip(d['公网IP'])
            d['源端口'] = self.extra_port(d['源端口'])
            d['目的端口'] = self.extra_port(d['目的端口'])
            d['公网端口'] = self.extra_port(d['公网端口'])
            # 将包含当前行需求信息的字典添加到需求列表中
            requirements_list.append(d)

        # 调用 merge_entries 函数合并相同的需求条目
        self.logger.info(f"合并前需求列表: {requirements_list}")
        requirements_list = self.merge_entries(requirements_list)
        self.logger.info(f"合并后需求列表: {requirements_list}")

        # 返回客户信息字典和合并后的需求列表
        return customer, requirements_list

    def extra_ip(self, cell_value):
        """
        从字符串中提取所有 IPv4 地址和域名。

        使用正则表达式匹配字符串中的 IPv4 地址格式（xxx.xxx.xxx.xxx）和域名格式，
        提取所有非重叠的匹配结果并返回。

        :param cell_value: 待提取 IP 地址或域名的字符串（通常来自 Excel 单元格值）
        :return: 包含所有提取到的 IPv4 地址和域名的列表；若无匹配则返回空列表
        """
        if isinstance(cell_value, str) and cell_value.lower() == "any":
            return ["Any"]
        elif cell_value in ("", r"/", None):
            return ["Any"]

        try:
            # 尝试解析为 IP 地址或网络
            if ipaddress.ip_network(cell_value) or ipaddress.ip_address(cell_value):
                return [cell_value]
        except (ValueError, TypeError):
            pass

        # 定义 IP 和域名的正则表达式
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        # 域名正则表达式，支持字母、数字、连字符和多级域名
        domain_pattern = r'\b(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}\b'

        # 确保输入是字符串类型
        if not isinstance(cell_value, str):
            cell_value = str(cell_value)

        # # 新增：去除斜杠"/"及其后的内容
        # cell_value = cell_value.split('/')[0]

        # 提取 IP 地址和域名
        ips = re.findall(ip_pattern, cell_value)
        domains = re.findall(domain_pattern, cell_value)

        # 合并结果并去重
        result = list(set(ips + domains))

        return result if result else []

    def extra_port(self, cell_value):
        """
        从字符串中提取所有端口号。

        使用正则表达式匹配字符串中的端口号格式（xxxx），
        提取所有非重叠的匹配结果并返回。

        :param cell_value: 待提取端口号的字符串（通常来自 Excel 单元格值）
        :return: 包含所有提取到的端口号的列表；若无匹配则返回空列表
        """
        if isinstance(cell_value, int):
            return [cell_value]
        if not isinstance(cell_value, str):
            return []
        ports = []
        # 匹配两种模式：数字范围(如123-456)或单个数字(如789)
        port_patterns = re.findall(r'(\d+-\d+|\d+)', cell_value)
        for pattern in port_patterns:
            if '-' in pattern:
                start, end = map(int, pattern.split('-'))
                ports.extend(range(start, end + 1))
            else:
                ports.append(int(pattern))
        return ports

    def _get_merge_key_strategy(self, entry):
        """获取合并键策略"""
        strategies = [
            {
                'condition': lambda e: e['源归属'] == "行业云" and e['目的归属'] == "互联网",
                'key_builder': lambda e: (e['操作'], tuple(e['公网IP']))
            },
            {
                'condition': lambda e: e['源归属'] == "互联网" and e['目的归属'] == "行业云",
                'key_builder': lambda e: (e['操作'], tuple(e['公网IP']), tuple(e['目的IP']))
            }
        ]

        for strategy in strategies:
            if strategy['condition'](entry):
                return strategy['key_builder'](entry)

        # 默认策略
        return (entry['源归属'], entry['目的归属'])

    def merge_entries(self, data):
        """
        将列表中“源IP”、“公网IP”、“操作”和“目的IP”相同的条目合并为一条数据，
        不同的源端口和目的端口的值会被放入列表。

        :param data: 包含多个字典条目的列表，每个字典代表一条记录
        :return: 合并后的条目列表
        """
        # 用于存储合并后的条目，键为 (源IP, 操作, 目的IP) 组成的元组
        merged_dict = {}
        for entry in data:
            # 根据源归属和目的归属确定合并键
            key = self._get_merge_key_strategy(entry)

            if key not in merged_dict:  # 如果键不存在创建新条目
                merged_dict[key] = entry.copy()
            else:
                # 如果键已存在合并端口列表
                for field in ['源端口', '目的端口', '公网端口']:
                    for port in entry[field]:
                        if port not in merged_dict[key][field]:
                            merged_dict[key][field].append(port)
                    merged_dict[key][field].sort()
                # 合并IP列表
                for field in ['源IP', '目的IP', '公网IP']:
                    if isinstance(entry[field], list):
                        for ip in entry[field]:
                            if ip not in merged_dict[key][field]:
                                merged_dict[key][field].append(ip)
                        merged_dict[key][field].sort()

        return list(merged_dict.values())

    def generate_title(self, customer_info, requirement_info):
        """生成网络策略变更申请标题
        
        根据客户组织名称生成标准格式的网络策略变更申请标题。
        标题格式为"{客户组织}网络策略变更申请"。
        
        :param customer_info: 客户信息字典，必须包含'organization'键
        :type customer_info: dict
        :param requirement_info: 需求信息字典
        :type requirement_info: dict
        :return: 格式化的变更申请标题字符串
        :rtype: str
        """
        title = f"{customer_info['organization']}网络策略变更申请"
        self.logger.info(f"生成变更标题: {title}")
        return title

    def generate_reason(self, customer_info, requirement_info):
        """生成网络策略变更原因说明

        使用Jinja2模板引擎渲染变更原因模板，生成详细的变更原因说明。
        模板文件为'reason.tpl'，包含客户信息和需求信息上下文。
        
        :param customer_info: 客户信息字典，包含组织名称、系统名称等信息
        :type customer_info: dict
        :param requirement_info: 需求信息字典，包含源IP、目的IP、端口等网络策略信息
        :type requirement_info: dict
        :return: 渲染后的变更原因说明文本
        :rtype: str
        """
        #
        st_info = self.get_source_target_config(requirement_info)
        source_target_pair = st_info['pair']

        # 加载模板
        env = Environment(loader=FileSystemLoader(self.templates_dir))
        template = env.get_template('reason.tpl')

        # 渲染模板
        reason = template.render(
            customer=customer_info,
            requirement=requirement_info,
            source_target_pair=source_target_pair
        )

        self.logger.info(f"生成变更原因:\n {reason}")
        return reason

    def generate_scheme(self, customer_info, requirement_info):
        """生成网络策略变更方案

        根据场景类型和需求信息生成相应的网络策略变更方案。
        支持多种场景：互联网->行业云、行业云->互联网、行业云->zltest、行业云->行业云。
        通过反射机制调用对应的上下文处理器，结合Jinja2模板渲染生成最终方案。
        
        :param customer_info: 客户信息字典，包含组织名称、系统名称等信息
        :type customer_info: dict
        :param requirement_info: 需求信息字典，包含场景、源IP、目的IP、端口等网络策略信息
        :type requirement_info: dict
        :return: 渲染后的变更方案文本
        :rtype: str
        :raises ValueError: 当场景信息缺失或不支持时抛出异常
        """

        scenario_type, config = self.get_scenario_config(requirement_info["源归属"], requirement_info["目的归属"])

        # 复用场景配置（多处使用）
        # 1. 构建通用上下文
        common_context = self._get_common_context(customer_info, requirement_info, scenario_type)
        self.logger.debug(f"通用上下文: {common_context}")
        # 调用特定场景上下文处理器
        # 查询config配置文件获取场景上下文处理器,通过反射器，调用类方法
        context_processor_str = config.get('context_processor', '')
        self.logger.debug(context_processor_str)  # 建议使用logger替代print进行调试

        # 修复反射调用：移除多余的self参数，调整参数顺序以匹配方法定义
        if context_processor_str and hasattr(self, context_processor_str):
            scenario_context = getattr(self, context_processor_str)(customer_info, requirement_info)
            self.logger.debug(f"调用场景上下文处理器: {scenario_context}")
        else:
            self.logger.error(f"上下文处理器不存在: {context_processor_str}")
            scenario_context = {}

        if not scenario_context:
            self.logger.warning("无法获取场景特定上下文，使用通用上下文")

        # 合并上下文并添加 customer 变量和vpn相关变量
        context = {**common_context, **scenario_context, 'customer': customer_info, 'requirement': requirement_info}
        self.logger.debug(f"渲染Jinja2模板context：{context}")

        # 3. 渲染模板并返回结果
        env = Environment(loader=FileSystemLoader(self.templates_dir))
        template = env.get_template(config["template"])  # 动态加载模板文件
        scheme = template.render(**context)

        # 4. 新增：检查现有配置并处理方案
        # processed_scheme = check_existing_config(raw_scheme, requirement, source_target_pair)
        # return processed_scheme

        self.logger.info(f"生成变更方案:\n {scheme}")
        return scheme

    def get_source_target_config(self, requirement):
        """获取源归属-目标归属组合的配置信息
        
        从需求信息中提取源归属和目标归属字段，确定场景类型并获取对应的场景配置。
        
        :param requirement: 需求信息字典，必须包含"源归属"和"目的归属"键
        :type requirement: dict
        :return: 包含源目标组合、场景配置和场景类型的字典
        :rtype: dict
        """
        source_attribution = requirement.get("源归属")
        target_attribution = requirement.get("目的归属")
        source_target_pair = (source_attribution, target_attribution)

        scenario_type, scenario_config = self.get_scenario_config(source_attribution, target_attribution)

        # 返回配置信息及源目标组合
        return {
            "pair": source_target_pair,
            "config": scenario_config,
            "scenario_type": scenario_type
        }

    def get_scenario_config(self, source, dest):
        """统一场景解析函数，避免重复条件判断
        
        根据源归属和目标归属的组合确定场景类型，并返回对应的场景配置信息。
        
        :param source: 源归属字符串，如"行业云"、"互联网"等
        :type source: str
        :param dest: 目标归属字符串，如"行业云"、"互联网"等
        :type dest: str
        :return: 包含场景类型和场景配置的元组
        :rtype: tuple
        :raises ValueError: 当场景组合未定义时抛出异常
        """
        scenario_mapping = {
            ("行业云", "行业云"): "industry_cloud_to_industry_cloud",
            ("行业云", "互联网"): "industry_cloud_to_internet",
            ("互联网", "行业云"): "internet_to_industry_cloud",
            ("行业云", "证联网测试网"): "industry_cloud_to_zltest",
        }

        key = (source, dest)
        if key not in scenario_mapping:
            raise ValueError(f"未定义的场景组合: {source} -> {dest}")

        scenario_type = scenario_mapping[key]
        return scenario_type, self.config['scenario_config'][scenario_type]

    def _get_common_context(self, customer, requirement, scenario_type):
        """获取通用上下文信息，支持多场景处理
        
        根据场景类型从配置中提取IP映射关系和端口字段，构建通用的上下文字典。
        该方法为不同场景提供统一的上下文数据结构，便于模板渲染使用。
        
        :param customer: 客户信息字典，包含组织名称等信息
        :type customer: dict
        :param requirement: 需求信息字典，包含IP地址、端口等网络策略信息
        :type requirement: dict
        :param scenario_type: 场景类型字符串，对应SCENARIO_CONFIG中的键
        :type scenario_type: str
        :return: 构建好的上下文字典，包含IP映射和操作类型等信息
        :rtype: dict
        :raises ValueError: 当场景类型不支持时抛出异常
        """
        # 验证场景类型是否支持
        if scenario_type not in self.config["scenario_config"]:
            raise ValueError(f"不支持的场景类型: {scenario_type}")

        # 获取场景配置
        config = self.config["scenario_config"][scenario_type]
        context = {}

        # 添加IP相关信息
        ip_mapping = config["ip_mapping"]
        for key, field in ip_mapping.items():
            context[key] = requirement.get(field, [])

        # 添加端口相关信息
        # port_info = {}
        # for port_field in config["port_fields"]:
        #     port_info[port_field] = requirement.get(port_field, [])
        # context["ports"] = port_info

        # 添加操作类型
        context["operation"] = requirement.get("操作", "允许")

        return context

    def _process_internet_industry_cloud_common(self, customer, requirement, ip_field):
        """互联网与行业云互访场景的通用处理函数
        
        处理互联网与行业云互访场景的通用逻辑，包括：
        1. 检索阿里云EIP地址
        2. 获取并筛选IPAM运营商地址
        3. 匹配深信服地址组
        
        :param customer: 客户信息字典，包含组织名称等信息
        :type customer: dict
        :param requirement: 需求信息字典，包含公网IP等网络策略信息
        :type requirement: dict
        :param ip_field: 用于检索EIP的IP字段名（"源IP"或"目的IP"）
        :type ip_field: str
        :return: 包含EIP、ISP地址信息和深信服地址组名称的上下文字典
        :rtype: dict
        """
        context = {}

        # -------------------------- 1. 检索阿里云EIP --------------------------
        # 创建AliYunProcessor实例并调用get_eip_address方法
        aliyun_processor = AliYunProcessor({})
        eip_ips = aliyun_processor.get_eip_address(customer, requirement, ip_field)
        if eip_ips is None:  # 处理EIP检索失败的情况
            context["eip_ips"] = "请检查EIP是否正确关联"
        else:
            context["eip_ips"] = eip_ips

        # -------------------------- 2. IPAM获取并筛选运营商地址 (名称：IP地址)--------------------------
        # 可抽象为通用函数的逻辑
        isp_ips = requirement['公网IP']
        if isp_ips == r"/":
            # 创建IpamProcessor实例来调用get_isp_address方法
            ipam_processor = IpamProcessor(None)  # 传入None作为占位符，实际使用时可能需要真实数据
            isp_ips = ipam_processor.get_isp_address(customer)
        isp_ips_info = {}
        for ip in isp_ips:
            if ip.startswith("1.192.170"):
                isp_ips_info[ip] = "电信"
            elif ip.startswith("218.28.38"):
                isp_ips_info[ip] = "联通"
            elif ip.startswith("39.165.248") or ip.startswith("2409:8745"):
                isp_ips_info[ip] = "移动"

        context["isp_ips_info"] = isp_ips_info

        self.logger.info(f"IPAM-客户{customer['organization']}的运营商地址为: {isp_ips_info}")

        # 查找运营商地址名称
        if '117.160.150.30' not in isp_ips and '218.28.144.146' not in isp_ips:
            # 创建IpamProcessor实例来调用get_isp_name_addresses方法
            ipam_processor = IpamProcessor(None)  # 传入None作为占位符，实际使用时可能需要真实数据
            isp_name_address = ipam_processor.get_isp_name_addresses(customer, isp_ips)

            if not isp_name_address:
                error_msg = f"IPAM-未找到匹配的运营商地址（公网IP: {isp_ips}，组织名: {customer['organization']}）,请核实地址并登记。"
                self.logger.error(error_msg)
            else:
                self.logger.info(f"IPAM-客户{customer['organization']}的运营商名称: {isp_name_address}")
                context["isp_ips_name"] = isp_name_address

        # -------------------------- 3. 深信服地址组匹配 --------------------------
        with open(os.path.join(jsons_dir, "ad", "custom_addresses.json"), "r", encoding="utf-8") as f:
            custom_addresses_json = json.load(f)

        ad_isp_address_name = None
        ad_eip_address_name = None
        for item in custom_addresses_json["items"]:
            item_ips = item["addresses"]
            if sorted(isp_ips) == sorted(item_ips):
                ad_isp_address_name = item["name"]
            if eip_ips[0] in item_ips:
                ad_eip_address_name = item["name"]
            if ad_isp_address_name and ad_eip_address_name:
                break

        if ad_isp_address_name:
            self.logger.info(f"深信服-ISP地址组: {ad_isp_address_name}，成员IP: {isp_ips}")
        if ad_eip_address_name:
            self.logger.info(f"深信服-EIP地址组: {ad_eip_address_name}，成员IP: {eip_ips}")

        if not ad_isp_address_name or not ad_eip_address_name:
            missing = []
            if not ad_isp_address_name:
                # 定义ad_isp_address_name
                missing.append(f"ISP地址列表 {isp_ips}")
                ad_isp_address_name = f"ISP-{customer['organization']}"
            if not ad_eip_address_name:
                missing.append(f"EIP地址列表 {eip_ips}")
                ad_eip_address_name = f"EIP-{customer['organization']}"
            self.logger.error(f"深信服-地址组匹配失败，"
                              f"缺失: {', '.join(missing)}, 请手动创建")

        context["ad_isp_address_name"] = ad_isp_address_name
        context["ad_eip_address_name"] = ad_eip_address_name

        return context

    def process_internet_to_industry_cloud(self, customer, requirement):
        """互联网→行业云场景的上下文处理器
        
        处理互联网到行业云的网络策略变更场景，调用通用处理函数处理EIP、ISP地址和深信服地址组匹配。
        
        :param customer: 客户信息字典，包含组织名称等信息
        :type customer: dict
        :param requirement: 需求信息字典，包含目的IP等网络策略信息
        :type requirement: dict
        :return: 包含EIP、ISP地址信息和深信服地址组名称的上下文字典，处理失败时返回None
        :rtype: dict or None
        """
        context = self._process_internet_industry_cloud_common(customer, requirement, ip_field="目的IP")
        if not context:
            return None
        return context

    def process_industry_cloud_to_internet(self, customer, requirement):
        """行业云→互联网场景的上下文处理器
        
        处理行业云到互联网的网络策略变更场景，调用通用处理函数处理EIP、ISP地址和深信服地址组匹配。
        
        :param customer: 客户信息字典，包含组织名称等信息
        :type customer: dict
        :param requirement: 需求信息字典，包含源IP等网络策略信息
        :type requirement: dict
        :return: 包含EIP、ISP地址信息和深信服地址组名称的上下文字典，处理失败时返回None
        :rtype: dict or None
        """
        context = self._process_internet_industry_cloud_common(customer, requirement, ip_field="源IP")
        self.logger.info(f"行业云→互联网场景的上下文处理器: {context}")
        if not context:
            return None
        return context

    def process_industry_cloud_to_zltest(self, customer, requirement):
        """行业云→zltest场景的上下文处理器
        
        处理行业云到zltest测试环境的网络策略变更场景，主要功能包括：
        1. 检索阿里云EIP地址
        2. 检索并匹配IPAM zlnet网段
        
        :param customer: 客户信息字典，包含组织名称等信息
        :type customer: dict
        :param requirement: 需求信息字典，包含源IP等网络策略信息
        :type requirement: dict
        :return: 包含EIP地址和匹配网段信息的上下文字典，处理失败时返回None
        :rtype: dict or None
        """
        SUBNETS_JSON_PATH = os.path.join(jsons_dir, "ipam", "subnets.json")
        subnets_data = _load_json_file(SUBNETS_JSON_PATH)

        context = {}

        # -------------------------- 1. 检索阿里云 EIP --------------------------
        # 调用抽象后的EIP检索函数
        # 创建AliYunProcessor实例来调用get_eip_address方法
        aliyun_processor = AliYunProcessor(None)  # 传入None作为占位符，实际使用时可能需要真实数据
        eip_ips = aliyun_processor.get_eip_address(customer, requirement, ip_field='源IP')
        if eip_ips is None:  # 处理EIP检索失败的情况
            return None
        context["eip_ips"] = eip_ips

        # -------------------------- 2. 检索IPAM zlnet网段 --------------------------
        # 调用抽象后的IPAM检索函数
        # 创建IpamProcessor实例来调用get_isp_address方法
        ipam_processor = IpamProcessor(subnets_data)
        ipam_cidrs = ipam_processor.get_ipam_cidrs(subnets_data)
        ipam_net = ipam_processor.get_longest_match_subnet(eip_ips, ipam_cidrs)
        if ipam_net is None:  # 处理IPAM检索失败的情况
            return None
        # 将返回的子网字符串转换为包含cidr字段的字典列表，以匹配模板期望的格式
        context["nets"] = [{"cidr": ipam_net}]
        #
        return context

    def process_industry_cloud_to_industry_cloud(self, customer, requirement):
        """行业云→行业云场景的上下文处理器
        
        处理行业云到行业云的网络策略变更场景，主要功能包括：
        1. 整合阿里云VPC信息与IPAM并网信息
        2. 合并相同网络属性的IP地址
        
        :param customer: 客户信息字典，包含组织名称等信息
        :type customer: dict
        :param requirement: 需求信息字典，包含源IP和目的IP等网络策略信息
        :type requirement: dict
        :return: 包含合并后的网络信息上下文，处理失败时返回包含错误信息的字典
        :rtype: dict
        """
        # 常量定义 - 提取配置参数，便于维护
        VPCS_JSON_PATH = os.path.join(jsons_dir, "ali_cloud", "vpcs.json")
        SUBNETS_JSON_PATH = os.path.join(jsons_dir, "ipam", "subnets.json")
        TARGET_CIDR = "172.27.0.0/16"

        try:
            # 1. 加载数据
            vpcs_data = _load_json_file(VPCS_JSON_PATH)
            subnets_data = _load_json_file(SUBNETS_JSON_PATH)

            # 2. 提取IP地址
            s_ips = requirement['源IP']
            d_ips = requirement['目的IP']
            all_ips = s_ips + d_ips

            # 3. 查找IP所在VPC
            combine_net_infos = get_combine_net_infos(all_ips, vpcs_data)
            self.logger.debug(f"combine_net_infos: {combine_net_infos}")

            # 4. 获取IPAM子网信息
            # 创建IpamProcessor实例并调用_get_ipam_cidrs方法
            ipam_processor = IpamProcessor(subnets_data)
            ipam_cidrs = ipam_processor.get_ipam_cidrs(subnets_data)
            self.logger.debug(f"ipam_all_cidrs: {ipam_cidrs}")

            # 5. 整合并网信息
            ipam_processor.enrich_combin_net_info_with_ipam(combine_net_infos, ipam_cidrs, TARGET_CIDR)

            # 6. 合并相同网络属性的IP
            merged_infos = ipam_processor.merge_similar_ips(combine_net_infos)
            context = {"combin_net_infos": merged_infos}

            return context

        except Exception as e:
            self.logger.error(f"处理网络信息时发生错误: {str(e)}", exc_info=True)
            return {"combin_net_infos": [], "error": str(e)}

    def _save_results_to_json(self, results, customer_info):
        """将处理结果保存为JSON文件
        
        将网络策略变更处理结果保存为JSON格式文件，文件名包含客户组织名称和当前日期。
        
        :param results: 处理结果数据，通常为包含网络策略信息的字典
        :type results: dict
        :param customer_info: 客户信息字典，必须包含"organization"键
        :type customer_info: dict
        """
        try:
            # 创建输出目录
            current_date = datetime.now().strftime("%Y%m%d")
            output_dir = os.path.join("/data/output/change_scripts", current_date)
            os.makedirs(output_dir, exist_ok=True)
            
            # 生成文件名
            org_name = customer_info.get("organization", "未知客户")
            filename = f"新增{org_name}-管理系统-访问策略变更.json"
            filepath = os.path.join(output_dir, filename)
            
            # 保存结果到JSON文件
            with open(filepath, 'w', encoding='utf-8-sig') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"结果已保存至: {filepath}")
        except Exception as e:
            self.logger.error(f"保存结果到JSON文件时发生错误: {str(e)}", exc_info=True)


class GetInformation:
    """获取信息类"""

    def __init__(self, config):
        """初始化方法，设置日志系统"""
        self.config = config
        self.logger = get_module_logger(__name__)

    def fetch_api_data(self, source_target_pair=None):
        """获取并保存各系统API数据（阿里云/深信服/山石）"""
        # 从配置文件加载场景配置
        scenario_fw_config = self.config.get('scenario_fw_config', {})
        default_config = self.config.get('default_fw_config', {
            "fw_url": "https://172.16.190.23",
            "sxf_save_dir": "ad/default",
            "fw_save_dir": "isp_fw/default"
        })

        # 转换配置键为元组格式，以便与现有代码兼容
        scenario_config = {}
        for key, value in scenario_fw_config.items():
            source_target = tuple(key.split('_'))
            scenario_config[source_target] = value

        # ======================== 通用数据获取（所有场景均执行）========================
        # 1. 阿里云API - 获取实例和网络接口信息
        from src.modules.apis.aliyun_api import AliYunApiClient
        from src.modules.apis.sangfor_ad import SangforApiClient
        from src.modules.apis.phpipam_api import PhpIpamApiClient
        from src.modules.apis.hillstone_fw import HillstoneApiClient
        # 1. 阿里云API - 获取实例和网络接口信息
        ali_api_client = AliYunApiClient()
        instances_info = ali_api_client.describe_instances()
        save_response_to_json(instances_info, "ali_cloud/instances.json", self.logger)

        network_interfaces_info = ali_api_client.describe_network_interfaces()
        save_response_to_json(network_interfaces_info, "ali_cloud/network_interfaces.json", self.logger)

        eips_info = ali_api_client.describe_eip_addresses()
        save_response_to_json(eips_info, "ali_cloud/eips.json", self.logger)

        # 2. IPAM API - 获取地址信息
        ipam_api_client = PhpIpamApiClient()
        addresses_info = ipam_api_client.get_addresses()
        save_response_to_json(addresses_info, "ipam/addresses.json", self.logger)

        subnets_info = ipam_api_client.get_subnets()
        save_response_to_json(subnets_info, "ipam/subnets.json", self.logger)

        # ======================== 场景特定数据获取（按source_target_pair执行）========================
        # 获取当前场景配置（优先场景 config，其次默认 config）
        current_config = scenario_config.get(source_target_pair,
                                             default_config) if source_target_pair else default_config

        # 3. 深信服AD API - 特定场景数据获取
        if source_target_pair == ("互联网", "行业云") or source_target_pair == ("行业云", "互联网"):
            sxf_api_client = SangforApiClient()
            dnat_info = sxf_api_client.get_dnat()
            save_response_to_json(dnat_info, f"{current_config['sxf_save_dir']}/dnat.json", self.logger)

            custom_address_info = sxf_api_client.get_custom_addresses()
            save_response_to_json(custom_address_info, f"{current_config['sxf_save_dir']}/custom_addresses.json",
                                  self.logger)

        # 4. 山石防火墙API - 按场景URL和路径获取数据
        hs_api_client = HillstoneApiClient(current_config["fw_url"])
        # 防火墙数据按场景分目录存储
        fw_save_dir = current_config["fw_save_dir"]
        # 场景一：互联网到行业云
        if source_target_pair == ("互联网", "行业云") or source_target_pair == ("行业云", "互联网"):
            hs_api_client = HillstoneApiClient(current_config["fw_url"])
            # 防火墙数据按场景分目录存储
            fw_save_dir = current_config["fw_save_dir"]

            policy_info = hs_api_client.get_policy()
            save_response_to_json(policy_info, f"{fw_save_dir}/policy.json", self.logger)

            addr_book_info = hs_api_client.get_addrbook()
            save_response_to_json(addr_book_info, f"{fw_save_dir}/addr_book.json", self.logger)

            service_ebook_info = hs_api_client.get_servicebook()
            save_response_to_json(service_ebook_info, f"{fw_save_dir}/service_book.json", self.logger)

            service_group_info = hs_api_client.get_servicegroup()
            save_response_to_json(service_group_info, f"{fw_save_dir}/service_group.json", self.logger)

        else:
            policy_info = hs_api_client.get_vsys_policy(current_config['fw_vsys'])
            save_response_to_json(policy_info, f"{fw_save_dir}/policy.json", self.logger)

            addr_book_info = hs_api_client.get_vsys_addrbook(current_config['fw_vsys'])
            save_response_to_json(addr_book_info, f"{fw_save_dir}/addr_book.json", self.logger)

            service_ebook_info = hs_api_client.get_vsys_servicebook(current_config['fw_vsys'])
            save_response_to_json(service_ebook_info, f"{fw_save_dir}/service_book.json", self.logger)


class IpamProcessor:
    """IPAM数据处理类"""

    def __init__(self, ipam_data):
        """初始化方法，设置日志系统"""
        self.ipam_data = ipam_data
        self.logger = get_module_logger(__name__)

    def get_ipam_cidrs(self, subnets_data):
        """从IPAM数据中提取子网和描述信息"""
        return {
            f"{subnet['subnet']}/{subnet['mask']}": subnet['description']
            for subnet in subnets_data
        }

    def get_longest_match_subnet(self, ips, subnets=None):
        """
        遵循最长匹配原则查找子网(支持IPv4/IPv6)
        :param ips: 要查询的IP地址列表
        :param subnets: 可选的自定义子网字典，格式{子网: 描述}
        :return: 最长匹配子网 或 None
        """
        try:
            ip_objs = [ipaddress.ip_address(ip) for ip in ips]
            target_subnets = subnets or self.ipam_data

            # 收集所有包含该IP的子网
            matched = []
            for subnet, desc in target_subnets.items():
                try:
                    net = ipaddress.ip_network(subnet, strict=False)
                    if any(ip_obj in net for ip_obj in ip_objs):
                        matched.append((net.prefixlen, subnet, desc))
                except ValueError:
                    continue

            # 按前缀长度降序排列
            if matched:
                matched.sort(reverse=True)
                return matched[0][1]  # 返回子网
            return None
        except ValueError:
            return None

    def get_isp_address(self, customer, ip=None):
        """获取客户的ISP地址列表

        :param customer: 客户信息字典
        :return: ISP地址列表
        """
        addresses_json_file = os.path.join(jsons_dir, "ipam", "addresses.json")
        with open(addresses_json_file, "r", encoding="utf-8") as f:
            addresses_json = json.load(f)

        # 1. 如果筛选主机名匹配的地址
        if ip:
            isp_addresses = []
            for item in addresses_json:
                # 显式处理hostname为None的情况：仅当hostname存在且非空时才检查前缀
                hostname = item.get('hostname')
                if hostname is not None and hostname.startswith(customer["organization"]):
                    isp_addresses.append(item['ip'])
            self.logger.info(f"IPAM-客户{customer['organization']}的ISP地址列表: {isp_addresses}")
            return isp_addresses

        # 模拟实现，返回空列表或默认值
        self.logger.warning("使用模拟实现的get_isp_address函数")
        return []

    def get_isp_name_addresses(self, customer, isp_ips):
        """获取ISP名称和地址的映射

        :param customer: 客户信息字典
        :param isp_ips: ISP IP地址列表
        :return: ISP名称和地址的映射字典
        """
        addresses_json_file = os.path.join(jsons_dir, "ipam", "addresses.json")
        with open(addresses_json_file, "r", encoding="utf-8") as f:
            addresses_json = json.load(f)

        # 匹配目标IP或主机名
        isp_addresses_name = set()
        for ip in isp_ips:
            hostname = jsonpath.jsonpath(addresses_json, f"$[?(@.ip == '{ip}')].hostname")
            if hostname:
                isp_addresses_name.add(hostname[0])
        if len(isp_addresses_name) > 1:
            self.logger.error(
                f"客户 {customer['organization']} 下的IP {isp_ips} 关联了多个主机名 {isp_addresses_name}，请检查IPAM配置。")

        return isp_addresses_name

    def enrich_combin_net_info_with_ipam(self, combin_net_infos, ipam_cidrs, target_cidr):
        """整合IPAM中的并网信息到网络信息中"""
        target_network = ipaddress.ip_network(target_cidr)

        for info in combin_net_infos.values():
            cidr = info['cidr']
            if not cidr or cidr not in ipam_cidrs:
                continue

            # 提取VLAN信息
            ipam_description = ipam_cidrs[cidr]
            vlan_match = re.search(r"vlan(\d+)", ipam_description)
            if not vlan_match:
                self.logger.warning(f"无法从描述中提取VLAN: {ipam_description}")
                continue

            combine_net_vlan = vlan_match.group(1)

            # 添加IPAM信息
            info.update({
                'vlan': combine_net_vlan,
                'description': re.search(r"(\S+)【", ipam_description).group(1),
                'vpn_instance': f"VPC-{combine_net_vlan}"
            })

            # 查找并网互联网段
            info['segment'], info['ip_a'], info['ip_b'], info['ip_c'] = self.find_internet_segment(
                ipam_cidrs, combine_net_vlan, target_network
            )

    def find_internet_segment(self, ipam_cidrs, vlan, target_network):
        """查找指定VLAN的并网对接网段"""
        for net, desc in ipam_cidrs.items():
            try:
                # 跳过IPv6
                if ipaddress.ip_network(net).version == 6:
                    continue

                # 检查是否包含目标VLAN且属于目标网段
                if desc and vlan in desc:
                    network = ipaddress.ip_network(net)
                    if network.subnet_of(target_network):
                        self.logger.info(f"行业云并网对接网段: {net}")
                        return net, self.get_ip_at_position(net, 3), self.get_ip_at_position(net,
                                                                                             4), self.get_ip_at_position(
                            net, 5)
            except ValueError as e:
                self.logger.warning(f"处理网段 {net} 时出错: {str(e)}")
                continue

        return None, None, None, None

    def merge_similar_ips(self, net_infos):
        """合并具有相同网络属性的IP地址"""
        merged = {}

        for ip, info in net_infos.items():
            # 使用网络信息作为合并键
            info_key = tuple(sorted(info.items()))

            if info_key not in merged:
                merged[info_key] = {**info, 'ips': []}
            merged[info_key]['ips'].append(ip)

        return list(merged.values())

    def get_ip_at_position(self, network_str, position, from_end=False):
        """
        获取网段中指定位置的IP地址

        参数:
            network_str: 网段字符串 (如 "192.168.1.0/24")
            position: 位置索引，从1开始
            from_end: 是否从网段末尾开始计数，默认为False（从开头计数）

        返回:
            指定位置的IP地址字符串，若位置无效则返回None
        """
        try:
            # 解析网段
            network = ipaddress.ip_network(network_str, strict=False)

            # 获取网段内所有可用IP地址列表
            # 注意：hosts()不包含网络地址和广播地址
            hosts = list(network.hosts())
            total_hosts = len(hosts)

            # 检查位置是否有效
            if position < 1 or position > total_hosts:
                print(f"无效位置！该网段共有 {total_hosts} 个可用IP地址")
                return None

            # 计算实际索引（列表从0开始）
            if from_end:
                index = total_hosts - position
            else:
                index = position - 1

            return str(hosts[index])

        except ValueError as e:
            print(f"网段解析错误: {e}")
            return None


class SangforAdProcessor:
    def __init__(self, sangfor_ad_data):
        self.sangfor_ad_data = sangfor_ad_data
        self.logger = get_module_logger(__name__)


class AliYunProcessor:
    def __init__(self, aliyun_data):
        self.aliyun_data = aliyun_data
        self.logger = get_module_logger(__name__)

    def get_eip_address(self, customer, requirement, ip_field):
        """从阿里云网络接口数据中检索EIP地址列表

        :param customer: 客户信息字典
        :param requirement: 需求信息字典
        :param ip_field: 用于检索EIP的IP字段名（"源IP"或"目的IP"）
        :return: EIP地址列表；若检索失败返回None
        """
        try:
            # 确保ip_field的值存在且是列表
            if ip_field not in requirement:
                self.logger.error(f"字段 {ip_field} 在需求信息中不存在")
                return None

            # 处理IP列表
            ip_list = requirement[ip_field]
            if isinstance(ip_list, str):
                ip_list = [ip_list]

            eip_addresses = []
            # 简单模拟EIP检索逻辑，避免依赖未定义的extract_jsonpath函数
            for ip in ip_list:
                if ipaddress.ip_address(ip) in ipaddress.ip_network("172.27.136.0/22"):
                    eip_addresses.append(ip)
                else:
                    # 通过JSONPath查询关联的公网EIP
                    eip_jsonpath = f'$.NetworkInterfaceSets.NetworkInterfaceSet[?(@.PrivateIpAddress=="{ip}")].AssociatedPublicIp.PublicIpAddress'
                    # 使用相对路径而不是硬编码的绝对路径
                    network_interfaces_path = os.path.join(jsons_dir, "ali_cloud", "network_interfaces.json")
                    try:
                        with open(network_interfaces_path, "r", encoding="utf-8") as f:
                            network_interfaces_json = json.load(f)
                        eip = extract_jsonpath(network_interfaces_json, eip_jsonpath)
                        if eip is None:
                            self.logger.error(f"ECS实例IP {ip} 未关联公网EIP，请关联后重新执行。")
                            exit()
                        # 将检索到的EIP加入eips
                        eip_addresses.append(eip)
                    except FileNotFoundError:
                        self.logger.error(f"网络接口JSON文件未找到: {network_interfaces_path}")
                        return None
                    except json.JSONDecodeError:
                        self.logger.error(f"网络接口JSON文件解析错误: {network_interfaces_path}")
                        return None
                    self.logger.info(f"阿里云-云内IP： {ip} <-> EIP地址为：{eip}")
            return eip_addresses
        except Exception as e:
            self.logger.error(f"获取EIP地址时出错：{str(e)}")
            return []


def main():
    """
    主函数，用于直接运行模块时调用
    """
    # 定义日志
    logger = get_module_logger(__name__)
    # 加载配置

    config_manager = ConfigManager(config_path)
    config = config_manager.load_config()
    logger.info('配置加载完成')
    logger.debug(f"配置信息：{config}")
    logger.debug(f"配置场景：{config['scenario_config'].keys()}")

    # 向上回溯两级到项目根目录（根据实际情况调整）
    # for root, dirs, files in os.walk(os.path.join(project_root, 'data', 'input', 'order')):
    #     for file in files:
    #         if file.endswith('.xlsx'):
    #             order_path = os.path.join(root, file)
    #             logger.info(f'开始处理订单: {order_path}')
    #             #
    #             content = OrderProcessor(order_path, config, template_dir)
    #             # print(content)

    order_path = os.path.join(project_root, 'data', 'input', 'order',
                              '网络资源-信达期货-证联网测试网.xlsx')

    logger.info(f'开始处理订单: {order_path}')
    #
    OrderProcessor(order_path, config, template_dir)


if __name__ == '__main__':
    main()
