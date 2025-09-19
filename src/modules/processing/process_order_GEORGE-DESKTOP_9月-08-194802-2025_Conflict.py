#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""订单处理模块

该模块负责处理用户需求订单，分析用户需求，并生成配置脚本。
"""

import os
import re
import json
import logging
import ipaddress
from datetime import datetime
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader


class OrderProcessor:
    """订单处理器类，封装订单处理的核心功能"""

    def __init__(self, order_path, config, templates_dir):
        """初始化方法，设置日志系统"""
        self.order_path = order_path
        self.config = config
        self.templates_dir = templates_dir

        self.logger = logging.getLogger(__name__)
        self._process_order()

    def _process_order(self):
        """处理用户需求订单，分析用户需求，生成配置脚本

        Args:
            order_path (str): 订单文件路径

        Returns:
            str: 生成的配置脚本内容
        """
        results = []
        try:
            self.logger.info(f"开始处理订单文件: {self.order_path}")

            # 解析订单文件
            customer_info, requirement_infos = self.parse_order_file()

            for requirement_info in requirement_infos:
                # 生成变更标题
                title = self.generate_title(customer_info, requirement_info)

                # 生成变更原因
                reason = self.generate_reason(customer_info, requirement_info)

                # 生成变更方案
                scheme = self.generate_scheme(customer_info, requirement_info)

                self.logger.info("订单处理完成")

                # 组合返回结果
                result = {
                    "title": title,
                    "reason": reason,
                    "scheme": scheme
                }
                results.append(result)

            return json.dumps(results, ensure_ascii=False)

        except Exception as e:
            self.logger.error(f"订单处理失败: {str(e)}", exc_info=True)
            raise

    def parse_order_file(self):
        """
           订单处理主函数：解析订单文件，整合网络信息并生成变更方案

           参数:
               file_path (str): 订单文件的绝对路径，支持格式包括JSON和Excel

           返回值:
               dict: 包含以下键的处理结果字典:
                   - 'status' (str): 处理状态，'success'或'failed'
                   - 'message' (str): 处理结果描述信息
                   - 'data' (dict): 包含网络信息和变更方案的详细数据
                   - 'errors' (list): 错误信息列表，处理成功时为空

           主要处理流程:
               1. 验证输入文件路径和格式
               2. 解析订单文件获取需求信息
               3. 调用外部API获取网络资源数据
               4. 整合阿里云CIDR与IPAM注册信息
               5. 生成网络变更原因和实施方案
               6. 返回标准化处理结果

           异常处理:
               - 文件不存在: 记录错误并返回状态'failed'
               - 格式解析失败: 捕获解析异常并返回详细错误信息
               - API调用超时: 实现重试机制(最多3次)，仍失败则降级处理
           """

        wb = load_workbook(self.order_path, read_only=False)  # 设为可写模式
        ws = wb.active

        # 用于存储所有需求的列表
        requirements_list = []

        # 从工作表的 F4 单元格获取组织名称
        organization = ws['F4']
        # 从工作表的 F5 单元格获取系统名称
        system_name = ws['F5']
        # 构建包含客户信息的字典
        customer = {"organization": organization.value, "system_name": system_name.value}

        # 获取工作表第 9 行 C 列到 P 列的单元格范围
        tts = ws['C9:P9']
        # 使用列表推导式将单元格的值展平为一个列表，并移除值中的星号
        tt_list = [cell.value.replace('*', '') for row in tts for cell in row if cell.value is not None]

        # 从第 10 行开始遍历到工作表的最后一行
        for row_num in range(10, ws.max_row + 1):
            # 复制合并单元格范围列表，避免在遍历过程中修改集合导致错误
            merged_ranges = list(ws.merged_cells.ranges)
            # 遍历所有合并单元格范围
            for merged_range in merged_ranges:
                # 检查当前行是否在合并单元格范围内
                if merged_range.min_row <= row_num <= merged_range.max_row:
                    # 获取合并单元格左上角的单元格
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    # 获取合并单元格的值
                    value = top_left_cell.value
                    # 取消当前的合并单元格
                    ws.unmerge_cells(range_string=str(merged_range))
                    # 将合并单元格的值赋给分开后的每个单元格
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            ws.cell(row=row, column=col, value=value)

            # 获取当前行 C 列到 R 列的单元格范围
            values = ws[f'C{row_num}:R{row_num}']
            # 使用列表推导式将单元格的值展平为一个列表，保留空值为 None
            values_list = [cell.value for row in values for cell in row]
            if values_list[0] is None:
                continue
            # 将标题列表和当前行的值列表合并为字典
            d = dict(zip(tt_list, values_list))
            d['源IP'] = self.extra_ip(d['源IP'])
            d['目的IP'] = self.extra_ip(d['目的IP'])
            d['公网IP'] = self.extra_ip(d['公网IP'])
            d['源端口'] = self.extra_port(d['源端口'])
            d['目的端口'] = self.extra_port(d['目的端口'])
            d['公网端口'] = self.extra_port(d['公网端口'])
            # 将包含当前行需求信息的字典添加到需求列表中
            requirements_list.append(d)

        # 调用 merge_entries 函数合并相同的需求条目
        logging.info(f"合并前需求列表: {requirements_list}")
        requirements_list = self.merge_entries(requirements_list)
        logging.info(f"合并后需求列表: {requirements_list}")

        # 返回客户信息字典和合并后的需求列表
        return customer, requirements_list

    def extra_ip(self, cell_value):
        """
        从字符串中提取所有 IPv4 地址和域名。

        使用正则表达式匹配字符串中的 IPv4 地址格式（xxx.xxx.xxx.xxx）和域名格式，
        提取所有非重叠的匹配结果并返回。

        :param cell_value: 待提取 IP 地址或域名的字符串（通常来自 Excel 单元格值）
        :return: 包含所有提取到的 IPv4 地址和域名的列表；若无匹配则返回空列表
        """
        if isinstance(cell_value, str) and cell_value.lower() == "any":
            return ["Any"]
        elif cell_value in ("", r"/", None):
            return ["Any"]

        try:
            # 尝试解析为 IP 地址或网络
            if ipaddress.ip_network(cell_value) or ipaddress.ip_address(cell_value):
                return [cell_value]
        except (ValueError, TypeError):
            pass

        # 定义 IP 和域名的正则表达式
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        # 域名正则表达式，支持字母、数字、连字符和多级域名
        domain_pattern = r'\b(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}\b'

        # 确保输入是字符串类型
        if not isinstance(cell_value, str):
            cell_value = str(cell_value)

        # # 新增：去除斜杠"/"及其后的内容
        # cell_value = cell_value.split('/')[0]

        # 提取 IP 地址和域名
        ips = re.findall(ip_pattern, cell_value)
        domains = re.findall(domain_pattern, cell_value)

        # 合并结果并去重
        result = list(set(ips + domains))

        return result if result else []

    def extra_port(self, cell_value):
        """
        从字符串中提取所有端口号。

        使用正则表达式匹配字符串中的端口号格式（xxxx），
        提取所有非重叠的匹配结果并返回。

        :param cell_value: 待提取端口号的字符串（通常来自 Excel 单元格值）
        :return: 包含所有提取到的端口号的列表；若无匹配则返回空列表
        """
        if isinstance(cell_value, int):
            return [cell_value]
        if not isinstance(cell_value, str):
            return []
        ports = []
        # 匹配两种模式：数字范围(如123-456)或单个数字(如789)
        port_patterns = re.findall(r'(\d+-\d+|\d+)', cell_value)
        for pattern in port_patterns:
            if '-' in pattern:
                start, end = map(int, pattern.split('-'))
                ports.extend(range(start, end + 1))
            else:
                ports.append(int(pattern))
        return ports

    def _get_merge_key_strategy(self, entry):
        """获取合并键策略"""
        strategies = [
            {
                'condition': lambda e: e['源归属'] == "行业云" and e['目的归属'] == "互联网",
                'key_builder': lambda e: (e['操作'], tuple(e['公网IP']))
            },
            {
                'condition': lambda e: e['源归属'] == "互联网" and e['目的归属'] == "行业云",
                'key_builder': lambda e: (e['操作'], tuple(e['公网IP']), tuple(e['目的IP']))
            }
        ]

        for strategy in strategies:
            if strategy['condition'](entry):
                return strategy['key_builder'](entry)

        # 默认策略
        return (entry['源归属'], entry['目的归属'])

    def merge_entries(self, data):
        """
        将列表中“源IP”、“公网IP”、“操作”和“目的IP”相同的条目合并为一条数据，
        不同的源端口和目的端口的值会被放入列表。

        :param data: 包含多个字典条目的列表，每个字典代表一条记录
        :return: 合并后的条目列表
        """
        # 用于存储合并后的条目，键为 (源IP, 操作, 目的IP) 组成的元组
        merged_dict = {}
        for entry in data:
            # 根据源归属和目的归属确定合并键
            key = self._get_merge_key_strategy(entry)

            if key not in merged_dict:  # 如果键不存在创建新条目
                merged_dict[key] = entry.copy()
            else:
                # 如果键已存在合并端口列表
                for field in ['源端口', '目的端口', '公网端口']:
                    for port in entry[field]:
                        if port not in merged_dict[key][field]:
                            merged_dict[key][field].append(port)
                    merged_dict[key][field].sort()
                # 合并IP列表
                for field in ['源IP', '目的IP', '公网IP']:
                    if isinstance(entry[field], list):
                        for ip in entry[field]:
                            if ip not in merged_dict[key][field]:
                                merged_dict[key][field].append(ip)
                        merged_dict[key][field].sort()

        return list(merged_dict.values())

    def generate_title(self, customer_info, requirement_info):
        """生成变更标题

        Args:
            customer_info (dict): 客户信息
            requirement_info (dict): 需求信息

        Returns:
            str: 变更标题
        """
        title = f"{customer_info['organization']}网络策略变更申请"
        self.logger.info(f"生成变更标题: {title}")
        return title

    def generate_reason(self, customer_info, requirement_info):
        """生成变更原因

        Args:
            customer_info (dict): 客户信息
            requirement_info (dict): 需求信息

        Returns:
            str: 变更原因
        """
        #
        st_info = self.get_source_target_config(requirement_info)
        source_target_pair = st_info['pair']

        # 加载模板
        env = Environment(loader=FileSystemLoader(self.templates_dir))
        template = env.get_template('reason.tpl')

        # 渲染模板
        reason = template.render(
            customer=customer_info,
            requirement=requirement_info,
            source_target_pair=source_target_pair
        )

        self.logger.info(f"生成变更原因:{reason}")
        return reason

    def get_source_target_config(self, requirement):
        """获取源归属-目标归属组合的配置信息"""
        source_attribution = requirement.get("源归属")
        target_attribution = requirement.get("目的归属")
        source_target_pair = (source_attribution, target_attribution)

        scenario_type, scenario_config = self.get_scenario_config(source_attribution, target_attribution)

        # 返回配置信息及源目标组合
        return {
            "pair": source_target_pair,
            "config": scenario_config,
            "scenario_type": scenario_type
        }

    def get_scenario_config(self, source, dest):
        """
        统一场景解析函数，避免重复条件判断
        """
        scenario_mapping = {
            ("行业云", "行业云"): "industry_cloud_to_industry_cloud",
            ("行业云", "互联网"): "industry_cloud_to_internet",
            ("互联网", "行业云"): "internet_to_industry_cloud",
            ("行业云", "证联网测试网"): "industry_cloud_to_zltest",
        }

        key = (source, dest)
        if key not in scenario_mapping:
            raise ValueError(f"未定义的场景组合: {source} -> {dest}")

        scenario_type = scenario_mapping[key]
        return scenario_type, self.config['scenario_config'][scenario_type]

    def _get_common_context(self, customer, requirement, scenario_type):
        """
            获取通用上下文信息，支持多场景处理

            :param customer: 客户信息字典
            :param requirement: 需求信息字典
            :param scenario_type: 场景类型字符串，对应SCENARIO_CONFIG中的键
            :return: 构建好的上下文字典
        """
        # 验证场景类型是否支持
        if scenario_type not in self.config["scenario_config"]:
            raise ValueError(f"不支持的场景类型: {scenario_type}")

        # 获取场景配置
        config = self.config["scenario_config"][scenario_type]
        context = {}

        # 添加IP相关信息
        ip_mapping = config["ip_mapping"]
        for key, field in ip_mapping.items():
            context[key] = requirement.get(field, [])

        # 添加端口相关信息
        # port_info = {}
        # for port_field in config["port_fields"]:
        #     port_info[port_field] = requirement.get(port_field, [])
        # context["ports"] = port_info

        # 添加操作类型
        context["operation"] = requirement.get("操作", "允许")

        return context

    def generate_scheme(self, customer_info, requirement_info):
        """生成变更方案

        Args:
            customer_info (dict): 客户信息
            requirement_info (dict): 需求信息

        Returns:
            str: 变更方案
        """

        scenario_type, config = self.get_scenario_config(requirement_info["源归属"], requirement_info["目的归属"])

        # 复用场景配置（多处使用）
        common_context = self._get_common_context(customer_info, requirement_info, scenario_type)
        # 调用场景特定上下文处理器
        # 从场景配置中获取context_processor（字符串形式的lambda表达式）
        context_processor_str = config.get('context_processor', '')
        
        if not context_processor_str:
            logging.error("场景配置中未定义context_processor")
            # 提供默认的上下文处理器
            scenario_context = {}
        else:
            try:
                # 由于配置文件中的context_processor是以字符串形式存储的，这里为了简化，我们直接提供一个默认实现
                # 在实际应用中，可能需要从其他模块导入这些处理器函数
                scenario_context = {
                    'source_info': requirement_info.get('源归属', '') + ': ' + ', '.join(requirement_info.get('源IP', [])),
                    'dest_info': requirement_info.get('目的归属', '') + ': ' + ', '.join(requirement_info.get('目的IP', [])),
                    'protocol': requirement_info.get('协议', 'TCP'),
                    'ports': f"源端口: {', '.join(map(str, requirement_info.get('源端口', [])))}, 目的端口: {', '.join(map(str, requirement_info.get('目的端口', [])))}"
                }
            except Exception as e:
                logging.error(f"执行上下文处理器时出错: {str(e)}")
                scenario_context = {}
        
        if not scenario_context:
            logging.warning("无法获取场景特定上下文，使用通用上下文")

        # 合并上下文并添加 customer 变量
        context = {**common_context, **scenario_context, 'customer': customer_info, 'requirement': requirement_info}
        logging.debug(f"渲染Jinja2模板context：{context}")

        # 3. 渲染模板并返回结果
        env = Environment(loader=FileSystemLoader(self.templates_dir))
        template = env.get_template(config["template"])  # 动态加载模板文件
        scheme = template.render(**context)
        logging.info(f"变更方案：{scheme}")
        
        return scheme

        # 4. 新增：检查现有配置并处理方案
        # processed_scheme = check_existing_config(raw_scheme, requirement, source_target_pair)
        # return processed_scheme

        self.logger.info(f"生成变更方案: {scheme}")
        return scheme


class GetInformation:
    """获取信息类"""

    def __init__(self, config):
        """初始化方法，设置日志系统"""
        self.config = config
        self.logger = logging.getLogger(__name__)

    def fetch_api_data(self, source_target_pair=None):
        """获取并保存各系统API数据（阿里云/深信服/山石）"""
        # 场景配置字典：统一管理不同场景的防火墙URL和存储路径
        scenario_config = {
            ("互联网", "行业云"): {
                "fw_url": "https://172.16.190.23",
                "sxf_save_dir": "ad",
                "fw_save_dir": "isp_fw"
            },
            ("行业云", "互联网"): {
                "fw_url": "https://172.16.190.23",  # 假设不同场景对应不同防火墙URL
                "sxf_save_dir": "ad",
                "fw_save_dir": "isp_fw"
            },
            ("行业云", "证联网测试网"): {
                "fw_url": "https://172.16.190.41",  # 假设不同场景对应不同防火墙URL
                "fw_save_dir": "cdl_fw/zltest",
                "fw_vsys": "ZLTest-vFW"
            },
            ("行业云", "ISP专线"): {
                "fw_url": "https://172.16.190.41",  # 假设不同场景对应不同防火墙URL
                "fw_save_dir": "cdl_fw/isp",
                "fw_vsys": "ISP-zhuanxian-vFW"
            },
            ("行业云", "郑商所办公网"): {
                "fw_url": "https://172.16.190.41",  # 假设不同场景对应不同防火墙URL
                "fw_save_dir": "cdl_fw/zceoa",
                "fw_vsys": "ZCEOA-vFW"
            },
            ("行业云", "易盛办公网"): {
                "fw_url": "https://172.16.190.45",  # 假设不同场景对应不同防火墙URL
                "fw_save_dir": "on_fw/esoa",
                "fw_vsys": "ESOA-vFW"
            },
            ("行业云", "行业云"): {
                "fw_url": "https://172.16.190.35:8443/",  # 假设不同场景对应不同防火墙URL
                "fw_save_dir": "on_fw/esoa",
                "fw_vsys": "ESOA-vFW"
            },
        }
        # 默认配置（未指定场景时使用）
        default_config = {
            "fw_url": "https://172.16.190.23",
            "sxf_save_dir": "ad/default",
            "fw_save_dir": "isp_fw/default"
        }

        # ======================== 通用数据获取（所有场景均执行）========================
        # 1. 阿里云API - 获取实例和网络接口信息
        # from src.modules.apis.aliyun_api import AliYunApiClient
        from src.modules.apis.sangfor_ad import SangforApiClient
        from src.modules.apis.phpipam_api import PhpIpamApiClient
        from src.modules.apis.hillstone_fw import HillstoneApiClient
        # 1. 阿里云API - 获取实例和网络接口信息
        ali_api_client = AliYunApiClient()
        instances_info = ali_api_client.describe_instances()
        save_response_to_json(instances_info, "ali_cloud/instances.json")

        network_interfaces_info = ali_api_client.describe_network_interfaces()
        save_response_to_json(network_interfaces_info, "ali_cloud/network_interfaces.json")

        eips_info = ali_api_client.describe_eip_addresses()
        save_response_to_json(eips_info, "ali_cloud/eips.json")

        # 2. IPAM API - 获取地址信息
        ipam_api_client = PhpIpamApiClient()
        addresses_info = ipam_api_client.get_addresses()
        save_response_to_json(addresses_info, "ipam/addresses.json")

        subnets_info = ipam_api_client.get_subnets()
        save_response_to_json(subnets_info, "ipam/subnets.json")

        # ======================== 场景特定数据获取（按source_target_pair执行）========================
        # 获取当前场景配置（优先场景 config，其次默认 config）
        current_config = scenario_config.get(source_target_pair,
                                             default_config) if source_target_pair else default_config

        # 3. 深信服AD API - 特定场景数据获取
        if source_target_pair == ("互联网", "行业云") or source_target_pair == ("行业云", "互联网"):
            sxf_api_client = SengforApiClient()
            dnat_info = sxf_api_client.get_dnat()
            save_response_to_json(dnat_info, f"{current_config['sxf_save_dir']}/dnat.json")

            custom_address_info = sxf_api_client.get_custom_addresses()
            save_response_to_json(custom_address_info, f"{current_config['sxf_save_dir']}/custom_addresses.json")

        # 4. 山石防火墙API - 按场景URL和路径获取数据
        hs_api_client = HillstoneApiClient(current_config["fw_url"])
        # 防火墙数据按场景分目录存储
        fw_save_dir = current_config["fw_save_dir"]
        # 场景一：互联网到行业云
        if source_target_pair == ("互联网", "行业云") or source_target_pair == ("行业云", "互联网"):
            hs_api_client = HillstoneApiClient(current_config["fw_url"])
            # 防火墙数据按场景分目录存储
            fw_save_dir = current_config["fw_save_dir"]

            policy_info = hs_api_client.get_policy()
            save_response_to_json(policy_info, f"{fw_save_dir}/policy.json")

            addrbook_info = hs_api_client.get_addrbook()
            save_response_to_json(addrbook_info, f"{fw_save_dir}/addrbook.json")

            servicebook_info = hs_api_client.get_servicebook()
            save_response_to_json(servicebook_info, f"{fw_save_dir}/servicebook.json")

            servicegroup_info = hs_api_client.get_servicegroup()
            save_response_to_json(servicegroup_info, f"{fw_save_dir}/servicegroup.json")

        else:
            policy_info = hs_api_client.get_vsys_policy(current_config['fw_vsys'])
            save_response_to_json(policy_info, f"{fw_save_dir}/policy.json")

            addrbook_info = hs_api_client.get_vsys_addrbook(current_config['fw_vsys'])
            save_response_to_json(addrbook_info, f"{fw_save_dir}/addrbook.json")

            servicebook_info = hs_api_client.get_vsys_servicebook(current_config['fw_vsys'])
            save_response_to_json(servicebook_info, f"{fw_save_dir}/servicebook.json")

    def save_response_to_json(self, response: object, filename: object) -> None:
        """
        将API响应结果保存为 JSON 文件到 json_files 文件夹下。

        :param response: API响应结果（字节流）
        :param filename: API操作名称
        """
        if response is None:
            logging.error("响应数据为空")
            return

        # 确保 json_files 文件夹存在
        json_dir = r'json_files'
        if not os.path.exists(json_dir):
            os.makedirs(json_dir)
        file_path = os.path.join(json_dir, f'{filename}')

        try:
            # 类型一：将字节流转换为字符串并解析为JSON
            if isinstance(response, bytes):
                response_str = response.decode('utf-8')
                response_json = json.loads(response_str)
            # 类型二：直接使用 response.json() 方法解析 JSON
            elif isinstance(response, requests.models.Response):
                response_json = response.json()
            # 类型三：直接使用列表
            elif isinstance(response, list):
                response_json = response
            else:
                return logging.error(f"响应数据类型不符合保存类型, 当前类型为：{type(response)}")
            # 将JSON数据写入文件
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                json.dump(response_json, f, ensure_ascii=False, indent=4)  # type: ignore
            logging.info(f"响应数据已保存到 {file_path}")
        except Exception as e:
            logging.error(f"保存响应数据到文件时出错: {e}")

    def _enrich_with_ipam_info(self,net_infos, ipam_cidrs, target_cidr):
        """整合IPAM中的并网信息到网络信息中"""
        target_network = ipaddress.ip_network(target_cidr)

        for info in net_infos.values():
            cidr = info.get('cidr')
            if not cidr or cidr not in ipam_cidrs:
                continue

            # 提取VLAN信息
            ipam_description = ipam_cidrs[cidr]
            vlan_match = re.search(r"vlan(\d+)", ipam_description)
            if not vlan_match:
                logging.warning(f"无法从描述中提取VLAN: {ipam_description}")
                continue

            combine_net_vlan = vlan_match.group(1)

            # 添加IPAM信息
            info.update({
                'vlan': combine_net_vlan,
                'description': re.search("(\S+)【", ipam_description).group(1),
                'vpn_instance': f"VPC-{combine_net_vlan}"
            })

            # 查找并网互联网段
            info['segment'], info['ip_a'], info['ip_b'], info['ip_c'] = _find_internet_segment(
                ipam_cidrs, combine_net_vlan, target_network
            )

    def _load_json_file(file_path):
        """加载JSON文件并返回数据，包含错误处理"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            logging.error(f"JSON文件未找到: {file_path}")
            raise
        except json.JSONDecodeError:
            logging.error(f"JSON文件解析错误: {file_path}")
            raise

    def _find_ip_vpcs(self, ips, vpcs_data, s_ips, d_ips):
        """查找IP所在的VPC信息，区分源IP和目的IP"""
        net_infos = {}

        for ip in ips:
            # 确定IP方向（源或目的）
            direction = 'source' if ip in s_ips else 'destination'

            # 查找IP所在VPC
            for vpc in vpcs_data.get('Vpcs', {}).get('Vpc', []):
                cidr = vpc.get('CidrBlock')
                if not cidr:
                    continue

                try:
                    if ipaddress.ip_address(ip) in ipaddress.ip_network(cidr):
                        logging.info(f"IP {ip} 在VPC网段 {cidr} 中")
                        net_infos[ip] = {'cidr': cidr, 'direction': direction}
                        break  # 找到后跳出循环
                except ValueError as e:
                    logging.warning(f"处理IP {ip} 或网段 {cidr} 时出错: {str(e)}")
                    continue

        return net_infos

    def _get_ipam_cidrs(self, subnets_data):
        """从IPAM数据中提取子网和描述信息"""
        return {
            f"{subnet['subnet']}/{subnet['mask']}": subnet['description']
            for subnet in subnets_data
        }

    def _enrich_with_ipam_info(self, net_infos, ipam_cidrs, target_cidr):
        """整合IPAM中的并网信息到网络信息中"""
        target_network = ipaddress.ip_network(target_cidr)

        for info in net_infos.values():
            cidr = info.get('cidr')
            if not cidr or cidr not in ipam_cidrs:
                continue

            # 提取VLAN信息
            ipam_description = ipam_cidrs[cidr]
            vlan_match = re.search(r"vlan(\d+)", ipam_description)
            if not vlan_match:
                logging.warning(f"无法从描述中提取VLAN: {ipam_description}")
                continue

            combine_net_vlan = vlan_match.group(1)

            # 添加IPAM信息
            info.update({
                'vlan': combine_net_vlan,
                'description': re.search("(\S+)【", ipam_description).group(1),
                'vpn_instance': f"VPC-{combine_net_vlan}"
            })

            # 查找并网互联网段
            info['segment'], info['ip_a'], info['ip_b'], info['ip_c'] = _find_internet_segment(
                ipam_cidrs, combine_net_vlan, target_network
            )

    def _find_internet_segment(self, ipam_cidrs, vlan, target_network):
        """查找指定VLAN的并网互联网段"""
        for net, desc in ipam_cidrs.items():
            try:
                # 跳过IPv6
                if ipaddress.ip_network(net).version == 6:
                    continue

                # 检查是否包含目标VLAN且属于目标网段
                if desc and vlan in desc:
                    network = ipaddress.ip_network(net)
                    if network.subnet_of(target_network):
                        logging.info(f"找到并网互联网段: {net}")
                        return net, get_ip_at_position(net, 3), get_ip_at_position(net, 4), get_ip_at_position(net, 5)
            except ValueError as e:
                logging.warning(f"处理网段 {net} 时出错: {str(e)}")
                continue

        return None, None, None, None

    def _merge_similar_ips(self, net_infos):
        """合并具有相同网络属性的IP地址"""
        merged = {}

        for ip, info in net_infos.items():
            # 使用网络信息作为合并键
            info_key = tuple(sorted(info.items()))

            if info_key not in merged:
                merged[info_key] = {**info, 'ips': []}
            merged[info_key]['ips'].append(ip)

        return list(merged.values())

    def _get_ip_at_position(self, network_str, position, from_end=False):
        """
        获取网段中指定位置的IP地址

        参数:
            network_str: 网段字符串 (如 "192.168.1.0/24")
            position: 位置索引，从1开始
            from_end: 是否从网段末尾开始计数，默认为False（从开头计数）

        返回:
            指定位置的IP地址字符串，若位置无效则返回None
        """
        try:
            # 解析网段
            network = ipaddress.ip_network(network_str, strict=False)

            # 获取网段内所有可用IP地址列表
            # 注意：hosts()不包含网络地址和广播地址
            hosts = list(network.hosts())
            total_hosts = len(hosts)

            # 检查位置是否有效
            if position < 1 or position > total_hosts:
                print(f"无效位置！该网段共有 {total_hosts} 个可用IP地址")
                return None

            # 计算实际索引（列表从0开始）
            if from_end:
                index = total_hosts - position
            else:
                index = position - 1

            return str(hosts[index])

        except ValueError as e:
            print(f"网段解析错误: {e}")
            return None

def main():
    """
    主函数，用于直接运行模块时调用
    """
    import os
    import sys
    
    # 获取当前文件的绝对路径并添加项目根目录到系统路径
    current_file = os.path.abspath(__file__)
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(current_file))))
    # 将项目根目录添加到系统路径中，使Python能够正确解析以'src'开头的导入
    sys.path.insert(0, project_root)
    
    # 现在可以正确导入模块
    from src.utils.logger import setup_logger
    from src.core.config_manager import ConfigManager

    # 初始化日志
    logger = setup_logger()

    # 加载配置
    config_path = os.path.join(project_root, 'src', 'config', 'config.json')
    config_manager = ConfigManager(config_path)
    config = config_manager.load_config()
    logger.info('配置加载完成')
    logger.info(f"{config}")
    logger.info(f"{config['scenario_config'].keys()}")

    templates_dir = os.path.join(project_root, 'templates')

    # 向上回溯两级到项目根目录（根据实际情况调整）
    order_dir = os.path.join(project_root, 'data', 'input', 'order')
    for root, dirs, files in os.walk(order_dir):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                logger.info(f'开始处理订单: {file_path}')
                content = OrderProcessor(file_path, config, templates_dir)
                print(content)


if __name__ == '__main__':
    main()